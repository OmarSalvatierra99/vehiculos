"""
Utilidades y acceso a datos para Inventarios OFS.
"""

import hashlib
import logging
import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger("INVENTARIOS")


def _normalizar_header(valor: str) -> str:
    if not valor:
        return ""
    limpio = re.sub(r"[^A-Z0-9]", "", str(valor).upper())
    return limpio


def _normalizar_clave(valor: str) -> str:
    if not valor:
        return ""
    normalizado = unicodedata.normalize("NFKD", str(valor))
    ascii_txt = "".join(ch for ch in normalizado if not unicodedata.combining(ch))
    ascii_txt = ascii_txt.upper()
    ascii_txt = re.sub(r"[^A-Z0-9]+", "_", ascii_txt).strip("_")
    return ascii_txt



def _parse_date(valor: Optional[str]) -> Optional[str]:
    if not valor:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    valor = str(valor).strip()
    if not valor:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(valor, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _hoy_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _hash_password(clave: str) -> str:
    return hashlib.sha256(clave.encode()).hexdigest()


ENTES_MANUALES = [
    "PODER EJECUTIVO DEL ESTADO DE TLAXCALA",
    "DESPACHO DE LA GOBERNADORA",
    "SECRETARÍA DE LA FUNCIÓN PÚBLICA",
    "SECRETARÍA DE IMPULSO AGROPECUARIO",
    "COORDINACIÓN DE COMUNICACIÓN",
    "SECRETARÍA DE MEDIO AMBIENTE",
    "SECRETARÍA DE CULTURA",
    "SECRETARÍA DE LAS MUJERES",
    "SECRETARÍA DE ORDENAMIENTO TERRITORIAL Y VIVIENDA",
    "SECRETARÍA DE SEGURIDAD CIUDADANA",
    "COORDINACIÓN GENERAL DE PLANEACIÓN E INVERSIÓN",
    "SECRETARÍA DE BIENESTAR",
    "SECRETARÍA DE GOBIERNO",
    "SECRETARÍA DE TRABAJO Y COMPETITIVIDAD",
    "CONSEJERÍA JURÍDICA DEL EJECUTIVO",
    "COORDINACIÓN ESTATAL DE PROTECCIÓN CIVIL",
    "SECRETARIADO EJECUTIVO DEL SISTEMA ESTATAL DE SEGURIDAD PÚBLICA",
    "INSTITUTO TLAXCALTECA DE DESARROLLO TAURINO",
    "INSTITUTO TLAXCALTECA DE ASISTENCIA ESPECIALIZADA A LA SALUD",
    "COMISIÓN ESTATAL DE ARBITRAJE MÉDICO",
    "CASA DE LAS ARTESANÍAS DE TLAXCALA",
    "PROCURADURÍA DE PROTECCIÓN AL AMBIENTE DEL ESTADO DE TLAXCALA",
    "INSTITUTO DE FAUNA SILVESTRE PARA EL ESTADO DE TLAXCALA",
    "OFICIALÍA MAYOR DE GOBIERNO",
    "SECRETARÍA DE FINANZAS",
    "SECRETARÍA DE DESARROLLO ECONÓMICO",
    "SECRETARÍA DE TURISMO",
    "SECRETARÍA DE INFRAESTRUCTURA",
    "SECRETARÍA DE EDUCACIÓN PÚBLICA",
    "SECRETARÍA DE MOVILIDAD Y TRANSPORTE",
    "COORDINACIÓN DE RADIO, CINE Y TELEVISIÓN",
    "EL COLEGIO DE TLAXCALA, A.C.",
    "FIDEICOMISO DE LA CIUDAD INDUSTRIAL DE XICOTÉNCATL",
    "COMISIÓN EJECUTIVA DE ATENCIÓN A VÍCTIMAS DEL ESTADO DE TLAXCALA",
    "FONDO MACRO PARA EL DESARROLLO INTEGRAL DE TLAXCALA",
    "INSTITUTO DE CAPACITACIÓN PARA EL TRABAJO DEL ESTADO DE TLAXCALA",
    "INSTITUTO DE CATASTRO DEL ESTADO DE TLAXCALA",
    "INSTITUTO DEL DEPORTE DE TLAXCALA",
    "INSTITUTO TECNOLÓGICO SUPERIOR DE TLAXCO",
    "INSTITUTO TLAXCALTECA DE LA INFRAESTRUCTURA FÍSICA EDUCATIVA",
    "PODER LEGISLATIVO DEL ESTADO DE TLAXCALA",
    "INSTITUTO TLAXCALTECA DE LA JUVENTUD",
    "INSTITUTO TLAXCALTECA PARA LA EDUCACIÓN DE LOS ADULTOS",
    "ORGANISMO PÚBLICO DESCENTRALIZADO SALUD DE TLAXCALA",
    "PATRONATO CENTRO DE REHABILITACIÓN INTEGRAL Y ESCUELA EN TERAPIA FÍSICA Y REHABILITACIÓN",
    "PATRONATO “LA LIBERTAD CENTRO CULTURAL DE APIZACO”",
    "PENSIONES CIVILES DEL ESTADO DE TLAXCALA",
    "SISTEMA ESTATAL PARA EL DESARROLLO INTEGRAL DE LA FAMILIA",
    "UNIDAD DE SERVICIOS EDUCATIVOS DEL ESTADO DE TLAXCALA",
    "UNIVERSIDAD POLITÉCNICA DE TLAXCALA",
    "UNIVERSIDAD POLITÉCNICA DE TLAXCALA REGIÓN PONIENTE",
    "PODER JUDICIAL DEL ESTADO DE TLAXCALA",
    "UNIVERSIDAD TECNOLÓGICA DE TLAXCALA",
    "UNIVERSIDAD INTERCULTURAL DE TLAXCALA",
    "ARCHIVO GENERAL E HISTÓRICO DEL ESTADO DE TLAXCALA",
    "TRIBUNAL DE JUSTICIA ADMINISTRATIVA DEL ESTADO DE TLAXCALA",
    "UNIVERSIDAD AUTÓNOMA DE TLAXCALA",
    "COMISIÓN ESTATAL DE DERECHOS HUMANOS",
    "INSTITUTO TLAXCALTECA DE ELECCIONES",
    "INSTITUTO DE ACCESO A LA INFORMACIÓN PÚBLICA Y PROTECCIÓN DE DATOS PERSONALES DEL ESTADO DE TLAXCALA",
    "TRIBUNAL DE CONCILIACIÓN Y ARBITRAJE DEL ESTADO DE TLAXCALA",
    "TRIBUNAL ELECTORAL DE TLAXCALA",
    "CENTRO DE CONCILIACIÓN LABORAL DEL ESTADO DE TLAXCALA",
    "FISCALÍA GENERAL DE JUSTICIA DEL ESTADO DE TLAXCALA",
    "SECRETARÍA EJECUTIVA DEL SISTEMA ANTICORRUPCIÓN DEL ESTADO DE TLAXCALA",
    "PATRONATO PARA LAS EXPOSICIONES Y FERIAS EN LA CIUDAD DE TLAXCALA",
    "COMISIÓN ESTATAL DEL AGUA Y SANEAMIENTO DEL ESTADO DE TLAXCALA",
    "COLEGIO DE BACHILLERES DEL ESTADO DE TLAXCALA",
    "COLEGIO DE EDUCACIÓN PROFESIONAL TÉCNICA DEL ESTADO DE TLAXCALA",
    "COLEGIO DE ESTUDIOS CIENTÍFICOS Y TECNOLÓGICOS DEL ESTADO DE TLAXCALA",
    "CONSEJO ESTATAL DE POBLACIÓN",
    "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE HUAMANTLA",
    "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE APIZACO",
    "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE CHIAUTEMPAN",
    "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE ZACATELCO",
    "COMISIÓN DE POTABLE Y Y ALCANTARILLADO DEL MUNICIPIO TLAXCALA",
]

MUNICIPIOS_MANUALES = [
    "ACUAMANALA DE MIGUEL HIDALGO",
    "CONTLA DE JUAN CUAMATZI",
    "CUAPIAXTLA",
    "CUAXOMULCO",
    "EL CARMEN TEQUEXQUITLA",
    "EMILIANO ZAPATA",
    "ESPAÑITA",
    "HUAMANTLA",
    "HUEYOTLIPAN",
    "IXTACUIXTLA DE MARIANO MATAMOROS",
    "IXTENCO",
    "ATLTZAYANCA",
    "LA MAGDALENA TLALTELULCO",
    "LÁZARO CÁRDENAS",
    "MAZATECOCHCO DE JOSÉ MARÍA MORELOS",
    "MUÑOZ DE DOMINGO ARENAS",
    "NANACAMILPA DE MARIANO ARISTA",
    "NATIVITAS",
    "PANOTLA",
    "PAPALOTLA DE XICOHTÉNCATL",
    "SAN DAMIÁN TEXOLOC",
    "SAN FRANCISCO TETLANOHCAN",
    "AMAXAC DE GUERRERO",
    "SAN JERÓNIMO ZACUALPAN",
    "SAN JOSÉ TEACALCO",
    "SAN JUAN HUACTZINCO",
    "SAN LORENZO AXOCOMANITLA",
    "SAN LUCAS TECOPILCO",
    "SAN PABLO DEL MONTE",
    "SANCTÓRUM DE LÁZARO CÁRDENAS",
    "SANTA ANA NOPALUCAN",
    "SANTA APOLONIA TEACALCO",
    "SANTA CATARINA AYOMETLA",
    "APETATITLÁN DE ANTONIO CARVAJAL",
    "SANTA CRUZ QUILEHTLA",
    "SANTA CRUZ TLAXCALA",
    "SANTA ISABEL XILOXOXTLA",
    "TENANCINGO",
    "TEOLOCHOLCO",
    "TEPETITLA DE LARDIZÁBAL",
    "TEPEYANCO",
    "TERRENATE",
    "TETLA DE LA SOLIDARIDAD",
    "TETLATLAHUCA",
    "APIZACO",
    "TLAXCALA",
    "TLAXCO",
    "TOCATLÁN",
    "TOTOLAC",
    "TZOMPANTEPEC",
    "XALOZTOC",
    "XALTOCAN",
    "XICOHTZINCO",
    "YAUHQUEMEHCAN",
    "ZACATELCO",
    "ATLANGATEPEC",
    "ZITLALTÉPEC DE TRINIDAD SÁNCHEZ SANTOS",
    "BENITO JUÁREZ",
    "CALPULALPAN",
    "CHIAUTEMPAN",
]


@dataclass
class MovimientoRow:
    data: Dict
    alerta: bool = False


class DatabaseManager:
    def __init__(self, db_path: str, catalogos_dir: str):
        self.db_path = db_path
        self.catalogos_dir = Path(catalogos_dir)
        logger.info("Base de datos en uso: %s", Path(self.db_path).resolve())
        self._init_db()
        self._migrate_schema()
        self._ensure_usuarios_columns()
        self._ensure_movimientos_columns()
        self._ensure_prestamos_columns()
        self._seed_usuarios()
        self._seed_vehiculos()
        self._seed_usuarios_vehiculos()
        self._seed_responsables()
        self._seed_auditores()
        self._seed_responsables_auditores()
        self._seed_resguardantes()

    def _connect(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self):
        conn = self._connect()
        cur = conn.cursor()
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                usuario TEXT UNIQUE NOT NULL,
                clave TEXT NOT NULL,
                rol TEXT NOT NULL DEFAULT 'usuario',
                puesto TEXT DEFAULT '',
                entes TEXT NOT NULL DEFAULT 'TODOS',
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS usuarios_personal (
                resguardante_id INTEGER NOT NULL,
                personal_id INTEGER NOT NULL,
                PRIMARY KEY (resguardante_id, personal_id),
                FOREIGN KEY(resguardante_id) REFERENCES usuarios(id),
                FOREIGN KEY(personal_id) REFERENCES usuarios(id)
            );

            CREATE TABLE IF NOT EXISTS entes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clave TEXT UNIQUE NOT NULL,
                nombre TEXT NOT NULL,
                tipo TEXT NOT NULL,
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS movimientos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                folio TEXT UNIQUE NOT NULL,
                usuario_id INTEGER NOT NULL,
                ente_clave TEXT NOT NULL,
                fecha_solicitud TEXT NOT NULL,
                fecha_entrega TEXT,
                fecha_devolucion TEXT,
                cantidad INTEGER NOT NULL,
                receptor_nombre TEXT NOT NULL,
                firma_recepcion TEXT,
                devuelto INTEGER DEFAULT 0,
                observaciones TEXT,
                resguardante_nombre TEXT,
                resguardante_id INTEGER,
                placa_unidad TEXT,
                marca TEXT,
                modelo TEXT,
                responsable_vehiculo TEXT,
                vehiculo_id INTEGER,
                responsable_id INTEGER,
                no_pasajeros INTEGER,
                ruta_destino TEXT,
                motivo_salida TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
            );

            CREATE TABLE IF NOT EXISTS vehiculos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                placa TEXT UNIQUE NOT NULL,
                modelo TEXT NOT NULL,
                marca TEXT NOT NULL,
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS responsables (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS resguardantes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS auditores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS responsables_auditores (
                responsable_id INTEGER NOT NULL,
                auditor_id INTEGER NOT NULL,
                orden INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY (responsable_id, auditor_id),
                FOREIGN KEY(responsable_id) REFERENCES responsables(id),
                FOREIGN KEY(auditor_id) REFERENCES auditores(id)
            );

            CREATE TABLE IF NOT EXISTS movimientos_auditores (
                movimiento_id INTEGER NOT NULL,
                auditor_id INTEGER NOT NULL,
                PRIMARY KEY (movimiento_id, auditor_id),
                FOREIGN KEY(movimiento_id) REFERENCES movimientos(id),
                FOREIGN KEY(auditor_id) REFERENCES auditores(id)
            );

            CREATE TABLE IF NOT EXISTS movimientos_destinos (
                movimiento_id INTEGER NOT NULL,
                ente_clave TEXT NOT NULL,
                orden INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY (movimiento_id, ente_clave, orden),
                FOREIGN KEY(movimiento_id) REFERENCES movimientos(id),
                FOREIGN KEY(ente_clave) REFERENCES entes(clave)
            );


            CREATE TABLE IF NOT EXISTS movimientos_eventos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                movimiento_id INTEGER NOT NULL,
                usuario_id INTEGER,
                evento TEXT NOT NULL,
                fecha TEXT NOT NULL,
                notas TEXT,
                FOREIGN KEY(movimiento_id) REFERENCES movimientos(id)
            );

            CREATE TABLE IF NOT EXISTS usuarios_vehiculos (
                usuario_id INTEGER NOT NULL,
                vehiculo_id INTEGER NOT NULL,
                PRIMARY KEY (usuario_id, vehiculo_id),
                FOREIGN KEY(usuario_id) REFERENCES usuarios(id),
                FOREIGN KEY(vehiculo_id) REFERENCES vehiculos(id)
            );

            CREATE TABLE IF NOT EXISTS prestamos_vehiculos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                solicitante_id INTEGER NOT NULL,
                propietario_id INTEGER NOT NULL,
                vehiculo_id INTEGER NOT NULL,
                fecha_solicitud TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'PENDIENTE',
                notas TEXT,
                fechas_solicitadas TEXT,
                FOREIGN KEY(solicitante_id) REFERENCES usuarios(id),
                FOREIGN KEY(propietario_id) REFERENCES usuarios(id),
                FOREIGN KEY(vehiculo_id) REFERENCES vehiculos(id)
            );
        """)
        conn.commit()
        conn.close()

    def _migrate_schema(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(movimientos)")
        existentes = {row["name"] for row in cur.fetchall()}
        columnas_obsoletas = {"item_id", "no_inventario", "tipo_notificacion_id", "auditores_nombres"}
        if existentes & columnas_obsoletas:
            cur.executescript("""
                CREATE TABLE IF NOT EXISTS movimientos_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    folio TEXT UNIQUE NOT NULL,
                    usuario_id INTEGER NOT NULL,
                    ente_clave TEXT NOT NULL,
                    fecha_solicitud TEXT NOT NULL,
                    fecha_entrega TEXT,
                    fecha_devolucion TEXT,
                    cantidad INTEGER NOT NULL,
                    receptor_nombre TEXT NOT NULL,
                    firma_recepcion TEXT,
                    devuelto INTEGER DEFAULT 0,
                    observaciones TEXT,
                    resguardante_nombre TEXT,
                    resguardante_id INTEGER,
                    placa_unidad TEXT,
                    marca TEXT,
                    modelo TEXT,
                    responsable_vehiculo TEXT,
                    vehiculo_id INTEGER,
                    responsable_id INTEGER,
                    no_pasajeros INTEGER,
                    ruta_destino TEXT,
                    motivo_salida TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(usuario_id) REFERENCES usuarios(id),
                    FOREIGN KEY(vehiculo_id) REFERENCES vehiculos(id)
                );
            """)
            columnas_objetivo = [
                "id",
                "folio",
                "usuario_id",
                "ente_clave",
                "fecha_solicitud",
                "fecha_entrega",
                "fecha_devolucion",
                "cantidad",
                "receptor_nombre",
                "firma_recepcion",
                "devuelto",
                "observaciones",
                "resguardante_nombre",
                "resguardante_id",
                "placa_unidad",
                "marca",
                "modelo",
                "responsable_vehiculo",
                "vehiculo_id",
                "responsable_id",
                "no_pasajeros",
                "ruta_destino",
                "motivo_salida",
                "created_at",
            ]
            columnas_migradas = [col for col in columnas_objetivo if col in existentes]
            columnas_txt = ", ".join(columnas_migradas)
            cur.execute(f"""
                INSERT INTO movimientos_new ({columnas_txt})
                SELECT {columnas_txt}
                FROM movimientos
            """)
            cur.execute("DROP TABLE movimientos")
            cur.execute("ALTER TABLE movimientos_new RENAME TO movimientos")

        cur.execute("DROP TABLE IF EXISTS inventario_items")
        cur.execute("DROP TABLE IF EXISTS notificaciones")
        cur.execute("DROP TABLE IF EXISTS movimientos_pasajeros")
        conn.commit()
        conn.close()

    def _ensure_usuarios_columns(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(usuarios)")
        existentes = {row["name"] for row in cur.fetchall()}
        if "puesto" not in existentes:
            cur.execute("ALTER TABLE usuarios ADD COLUMN puesto TEXT DEFAULT ''")
        conn.commit()
        conn.close()

    def _ensure_movimientos_columns(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(movimientos)")
        existentes = {row["name"] for row in cur.fetchall()}
        columnas = [
            ("resguardante_nombre", "TEXT"),
            ("resguardante_id", "INTEGER"),
            ("placa_unidad", "TEXT"),
            ("marca", "TEXT"),
            ("modelo", "TEXT"),
            ("responsable_vehiculo", "TEXT"),
            ("vehiculo_id", "INTEGER"),
            ("responsable_id", "INTEGER"),
            ("no_pasajeros", "INTEGER"),
            ("ruta_destino", "TEXT"),
            ("motivo_salida", "TEXT"),
        ]
        for nombre, tipo in columnas:
            if nombre not in existentes:
                cur.execute(f"ALTER TABLE movimientos ADD COLUMN {nombre} {tipo}")
        conn.commit()
        conn.close()

    def _ensure_prestamos_columns(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(prestamos_vehiculos)")
        existentes = {row["name"] for row in cur.fetchall()}
        columnas = [
            ("responsable_id", "INTEGER"),
            ("responsable_nombre", "TEXT"),
            ("no_pasajeros", "INTEGER"),
            ("pasajeros_ids", "TEXT"),
            ("ruta_destino", "TEXT"),
            ("motivo_salida", "TEXT"),
            ("fechas_solicitadas", "TEXT"),
        ]
        for nombre, tipo in columnas:
            if nombre not in existentes:
                cur.execute(f"ALTER TABLE prestamos_vehiculos ADD COLUMN {nombre} {tipo}")
        conn.commit()
        conn.close()

    def _seed_usuarios(self):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM usuarios")
        total = cur.fetchone()[0]

        usuarios_path = self.catalogos_dir / "Usuarios_SASP_2025.xlsx"
        usuarios = []

        try:
            from openpyxl import load_workbook
        except ImportError:
            load_workbook = None

        if load_workbook and usuarios_path.exists():
            wb = load_workbook(usuarios_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(min_row=1, values_only=True)
            headers = next(rows, [])
            header_map = {
                _normalizar_header(h): idx
                for idx, h in enumerate(headers or [])
                if h
            }

            def get_val(row, keys):
                for key in keys:
                    for h, idx in header_map.items():
                        if key in h:
                            return row[idx]
                return None

            for row in rows:
                nombre = get_val(row, ["NOMBRE"])
                usuario = get_val(row, ["USUARIO"])
                clave = get_val(row, ["CLAVE", "PASSWORD"])
                entes = get_val(row, ["ENTES", "ENTE"])
                rol = get_val(row, ["ROL", "PERFIL"])
                puesto = get_val(row, ["PUESTO", "CARGO"])
                if not nombre or not usuario:
                    continue

                clave_txt = str(clave).strip() if clave else f"{usuario}2025"
                rol_txt = str(rol).strip().lower() if rol else "usuario"
                if "gestor" in rol_txt or "admin" in rol_txt or "monitor" in rol_txt:
                    rol_txt = "admin"
                elif "usuario" in rol_txt or "user" in rol_txt:
                    rol_txt = "user"
                elif rol_txt not in {"admin", "user"}:
                    rol_txt = "user"

                entes_txt = str(entes).strip().upper() if entes else "TODOS"
                usuarios.append((
                    str(nombre).strip(),
                    str(usuario).strip(),
                    _hash_password(clave_txt),
                    rol_txt,
                    str(puesto).strip() if puesto else "",
                    entes_txt,
                ))

        if total == 0 and not usuarios:
            usuarios = [
                ("C.P. Miguel Ángel Roldán Peña", "miguel", _hash_password("miguel2025"), "user", "", "TODOS"),
                ("C.P. Cristina Rosas de la Cruz", "cristina", _hash_password("cristina2025"), "user", "", "TODOS"),
                ("C.P. Ángel Flores Licona", "angel", _hash_password("angel2025"), "user", "", "TODOS"),
                ("C.P. Juan José Blanco Sánchez", "juan", _hash_password("juan2025"), "user", "", "TODOS"),
                ("Ing. Omar Alfredo Castro Orozco", "omar", _hash_password("omar2025"), "user", "", "TODOS"),
            ]
            logger.warning("Usuarios base creados; cambie las claves por seguridad.")

        if total == 0 and usuarios:
            cur.executemany(
                "INSERT INTO usuarios (nombre, usuario, clave, rol, puesto, entes) VALUES (?, ?, ?, ?, ?, ?)",
                usuarios,
            )

        usuarios_requeridos = [
            ("C.P. Miguel Ángel Roldán Peña", "miguel", "miguel2025", "user", ""),
            ("C.P. Cristina Rosas de la Cruz", "cristina", "cristina2025", "user", ""),
            ("C.P. Ángel Flores Licona", "angel", "angel2025", "user", ""),
            ("C.P. Juan José Blanco Sánchez", "juan", "juan2025", "user", ""),
            ("Ing. Omar Alfredo Castro Orozco", "omar", "omar2025", "user", ""),
        ]

        for nombre, usuario, clave_txt, rol_txt, puesto in usuarios_requeridos:
            cur.execute("""
                SELECT id
                FROM usuarios
                WHERE LOWER(usuario)=LOWER(?)
            """, (usuario,))
            if cur.fetchone():
                continue
            cur.execute(
                "INSERT INTO usuarios (nombre, usuario, clave, rol, puesto, entes) VALUES (?, ?, ?, ?, ?, ?)",
                (nombre, usuario, _hash_password(clave_txt), rol_txt, puesto, "TODOS"),
            )

        conn.commit()
        conn.close()

    # -------------------------------------------------------
    # Usuarios
    # -------------------------------------------------------
    def get_usuario(self, usuario: str, clave: str):
        if not usuario or not clave:
            return None
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre, usuario, clave, rol, entes
            FROM usuarios
            WHERE LOWER(usuario)=LOWER(?)
              AND activo=1
            LIMIT 1
        """, (usuario,))
        row = cur.fetchone()
        conn.close()
        if not row:
            return None
        if _hash_password(clave) != row["clave"]:
            return None
        entes = [e.strip().upper() for e in (row["entes"] or "").split(",") if e.strip()]
        return {
            "id": row["id"],
            "nombre": row["nombre"],
            "usuario": row["usuario"],
            "rol": row["rol"],
            "entes": entes or ["TODOS"],
        }

    # -------------------------------------------------------
    # Catálogos
    # -------------------------------------------------------
    def listar_entes(self) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT clave, nombre, tipo
            FROM entes
            WHERE activo=1
            ORDER BY nombre
        """)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def listar_usuarios(self, resguardante_id: Optional[int] = None) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        if resguardante_id:
            cur.execute("""
                SELECT usuario
                FROM usuarios
                WHERE id=?
            """, (resguardante_id,))
            row = cur.fetchone()
            if row and (row["usuario"] or "").lower() in {"luis", "odilia"}:
                resguardante_id = None

        if resguardante_id:
            cur.execute("""
                SELECT u.id, u.nombre, u.usuario, u.rol, u.puesto
                FROM usuarios u
                JOIN usuarios_personal up ON up.personal_id = u.id
                WHERE up.resguardante_id=?
                  AND u.activo=1
                  AND u.id != ?
                ORDER BY u.nombre
            """, (resguardante_id, resguardante_id))
        else:
            cur.execute("""
                SELECT id, nombre, usuario, rol, puesto
                FROM usuarios
                WHERE activo=1
                ORDER BY nombre
            """)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def listar_auditores_por_usuario(self, usuario_id: int) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT nombre
            FROM usuarios
            WHERE id=? AND activo=1
        """, (usuario_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return []

        responsable_nombre = row["nombre"]
        cur.execute("""
            SELECT a.id, a.nombre
            FROM responsables r
            JOIN responsables_auditores ra ON ra.responsable_id = r.id
            JOIN auditores a ON a.id = ra.auditor_id
            WHERE r.nombre=? AND a.activo=1
            ORDER BY ra.orden, a.nombre
        """, (responsable_nombre,))
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def listar_personal_resguardante(self, usuario_id: int) -> List[Dict]:
        auditores = self.listar_auditores_por_usuario(usuario_id)
        if not auditores:
            auditores = self.listar_auditores()
        return auditores

    def listar_resguardantes(self) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre
            FROM resguardantes
            WHERE activo=1
            ORDER BY nombre
        """)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def _seed_vehiculos(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM vehiculos")
        if cur.fetchone()[0] > 0:
            conn.close()
            return

        vehiculos = [
            ("XVZ-360-C", "Versa Sense", "NISSAN"),
            ("XVZ-373-C", "Versa Sense", "NISSAN"),
            ("XVZ-351-C", "Versa", "NISSAN"),
            ("XVZ-357-C", "Versa Sense", "NISSAN"),
            ("XVZ-358-C", "Versa Sense", "NISSAN"),
            ("XVZ-346-C", "Versa", "NISSAN"),
            ("XVZ-370-C", "Versa Sense", "NISSAN"),
            ("XVZ-356-C", "Versa Sense", "NISSAN"),
            ("XTR-479-E", "Aveo", "CHEVROLET"),
            ("XVZ-359-C", "Versa Sense", "NISSAN"),
            ("XVZ-371-C", "Versa Sense", "NISSAN"),
            ("XVZ-349-C", "Versa", "NISSAN"),
            ("XVZ-353-C", "Versa", "NISSAN"),
            ("XXK-741-D", "Gol Sedán", "VOLKSWAGEN"),
            ("XC-8407-C", "NP 300", "NISSAN"),
            ("XVZ-335-C", "Aveo", "CHEVROLET"),
        ]
        cur.executemany("""
            INSERT INTO vehiculos (placa, modelo, marca, activo)
            VALUES (?, ?, ?, 1)
        """, vehiculos)
        conn.commit()
        conn.close()

    def _seed_usuarios_vehiculos(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM usuarios_vehiculos")
        if cur.fetchone()[0] > 0:
            conn.close()
            return

        cur.execute("""
            SELECT id, usuario
            FROM usuarios
            WHERE activo=1
        """)
        usuarios = {row["usuario"].lower(): row["id"] for row in cur.fetchall()}

        cur.execute("""
            SELECT id, placa
            FROM vehiculos
            WHERE activo=1
        """)
        vehiculos = {row["placa"].upper(): row["id"] for row in cur.fetchall()}

        asignaciones = [
            ("cristina", ["XVZ-357-C", "XVZ-358-C", "XVZ-346-C"]),
            ("miguel", ["XVZ-360-C", "XVZ-373-C", "XVZ-351-C"]),
            ("omar", ["XVZ-353-C", "XXK-741-D", "XC-8407-C", "XVZ-335-C"]),
            ("angel", ["XVZ-370-C", "XVZ-356-C", "XTR-479-E"]),
            ("juan", ["XVZ-359-C", "XVZ-371-C", "XVZ-349-C"]),
        ]

        registros = []
        for usuario_key, placas in asignaciones:
            usuario_id = usuarios.get(usuario_key)
            if not usuario_id:
                continue
            for placa in placas:
                vehiculo_id = vehiculos.get(placa.upper())
                if vehiculo_id:
                    registros.append((usuario_id, vehiculo_id))

        if registros:
            cur.executemany("""
                INSERT OR IGNORE INTO usuarios_vehiculos (usuario_id, vehiculo_id)
                VALUES (?, ?)
            """, registros)

        conn.commit()
        conn.close()

    def _seed_responsables(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM responsables")
        if cur.fetchone()[0] > 0:
            conn.close()
            return

        responsables = [
            ("C.P. Miguel Ángel Roldán Peña",),
            ("C.P. Cristina Rosas de la Cruz",),
            ("C.P. Ángel Flores Licona",),
            ("C.P. Juan José Blanco Sánchez",),
            ("Ing. Omar Alfredo Castro Orozco",),
        ]
        cur.executemany("""
            INSERT INTO responsables (nombre, activo)
            VALUES (?, 1)
        """, responsables)
        conn.commit()
        conn.close()

    def _seed_resguardantes(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM resguardantes")
        if cur.fetchone()[0] > 0:
            conn.close()
            return

        resguardantes = [
            ("Luis Aguilar",),
            ("Monica Perez",),
            ("Daniel Ortiz",),
        ]
        cur.executemany("""
            INSERT INTO resguardantes (nombre, activo)
            VALUES (?, 1)
        """, resguardantes)
        conn.commit()
        conn.close()

    def _seed_auditores(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM auditores")
        if cur.fetchone()[0] > 0:
            conn.close()
            return

        auditores = [
            ("C.P. Yaneth Cruz George",),
            ("C.P. Julissa Karen Flores Pérez",),
            ("C.P. Margaret Michelle Pluma Meléndez",),
            ("C.P. Vanesa Angulo Ramírez",),
            ("C.P. Antonio Mastranzo Sánchez",),
            ("C.P. Gonzalo Flores Pérez",),
            ("Lic. Liliana Bonilla Montiel",),
            ("C.P. Paola Rodríguez Sánchez",),
            ("C.P. David Yair Juárez Zainos",),
            ("C.P. Mayra Ortega Campech",),
            ("C.P. Omar Romero Flores",),
            ("Lic. Iván Xahuentitla Domínguez",),
            ("C.P. María Fernanda Vázquez Ramírez",),
            ("C.P. Luis Enrique Velázquez López",),
            ("C.P. Beatriz Netzahualcóyotl Nava",),
            ("C.P. Patricia Romano López",),
            ("C.P. Diana Angélica Mendoza Cortés",),
            ("C.P. Gloria Arévalo Gutiérrez",),
            ("C.P. Eliazar Nava Nava",),
            ("C.P. Edgar Daniel Ordoñez Salinas",),
            ("C.P. Melina Flores Peña",),
            ("C.P. Roberto Sánchez Espinoza",),
            ("C.P. Jonathan Islas Sosa",),
            ("C.P. Isael López Cervantes",),
            ("C.P. Aranza Sánchez Trejo",),
            ("Arq. Ramos Martín Quiebras Techalotzi",),
            ("Arq. Jesús Coca López",),
            ("Arq. Juan Eduardo López García",),
            ("Arq. José Miguel Ángel Morales Vásquez",),
            ("Arq. Guadalupe Carrillo Raya",),
            ("Ing. Sergio Grajeda Avendaño",),
            ("Ing. Ubaldo Cuapio Cuapio",),
            ("Arq. Sergio David Jiménez Cuahutepitzi",),
            ("Ing. Alfonso Luis Vázquez Barrera",),
            ("Arq. Arely López Breton",),
            ("Ing. Yessica Flores Flores",),
            ("Arq. Damaris Pérez Cárcamo",),
            ("Arq. Jesús Alberto Islas López",),
            ("Arq. José Luis Ramírez Hernández",),
            ("Arq. Monserrat Moncayo Gómez",),
            ("Ing. Omar Alejandro Limón",),
            ("Arq. Yazmín Zárate Pérez",),
        ]
        cur.executemany("""
            INSERT INTO auditores (nombre, activo)
            VALUES (?, 1)
        """, auditores)
        conn.commit()
        conn.close()

    def _seed_responsables_auditores(self) -> None:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM responsables_auditores")
        if cur.fetchone()[0] > 0:
            conn.close()
            return

        cur.execute("SELECT id, nombre FROM responsables WHERE activo=1")
        responsables = {row["nombre"]: row["id"] for row in cur.fetchall()}
        cur.execute("SELECT id, nombre FROM auditores WHERE activo=1")
        auditores = {row["nombre"]: row["id"] for row in cur.fetchall()}

        grupos = [
            ("C.P. Miguel Ángel Roldán Peña", [
                "C.P. Yaneth Cruz George",
                "C.P. Julissa Karen Flores Pérez",
                "C.P. Margaret Michelle Pluma Meléndez",
                "C.P. Vanesa Angulo Ramírez",
                "C.P. Antonio Mastranzo Sánchez",
                "C.P. Gonzalo Flores Pérez",
            ]),
            ("C.P. Cristina Rosas de la Cruz", [
                "Lic. Liliana Bonilla Montiel",
                "C.P. Paola Rodríguez Sánchez",
                "C.P. David Yair Juárez Zainos",
                "C.P. Mayra Ortega Campech",
                "C.P. Omar Romero Flores",
                "Lic. Iván Xahuentitla Domínguez",
                "C.P. María Fernanda Vázquez Ramírez",
            ]),
            ("C.P. Ángel Flores Licona", [
                "C.P. Luis Enrique Velázquez López",
                "C.P. Beatriz Netzahualcóyotl Nava",
                "C.P. Patricia Romano López",
                "C.P. Diana Angélica Mendoza Cortés",
                "C.P. Gloria Arévalo Gutiérrez",
                "C.P. Eliazar Nava Nava",
            ]),
            ("C.P. Juan José Blanco Sánchez", [
                "C.P. Edgar Daniel Ordoñez Salinas",
                "C.P. Melina Flores Peña",
                "C.P. Roberto Sánchez Espinoza",
                "C.P. Jonathan Islas Sosa",
                "C.P. Isael López Cervantes",
                "C.P. Aranza Sánchez Trejo",
            ]),
            ("Ing. Omar Alfredo Castro Orozco", [
                "Arq. Ramos Martín Quiebras Techalotzi",
                "Arq. Jesús Coca López",
                "Arq. Juan Eduardo López García",
                "Arq. José Miguel Ángel Morales Vásquez",
                "Arq. Guadalupe Carrillo Raya",
                "Ing. Sergio Grajeda Avendaño",
                "Ing. Ubaldo Cuapio Cuapio",
                "Arq. Sergio David Jiménez Cuahutepitzi",
                "Ing. Alfonso Luis Vázquez Barrera",
                "Arq. Arely López Breton",
                "Ing. Yessica Flores Flores",
                "Arq. Damaris Pérez Cárcamo",
                "Arq. Jesús Alberto Islas López",
                "Arq. José Luis Ramírez Hernández",
                "Arq. Monserrat Moncayo Gómez",
                "Ing. Omar Alejandro Limón",
                "Arq. Yazmín Zárate Pérez",
            ]),
        ]

        registros = []
        for responsable_nombre, lista_auditores in grupos:
            responsable_id = responsables.get(responsable_nombre)
            if not responsable_id:
                continue
            for orden, auditor_nombre in enumerate(lista_auditores, start=1):
                auditor_id = auditores.get(auditor_nombre)
                if auditor_id:
                    registros.append((responsable_id, auditor_id, orden))

        if registros:
            cur.executemany("""
                INSERT OR IGNORE INTO responsables_auditores (responsable_id, auditor_id, orden)
                VALUES (?, ?, ?)
            """, registros)

        conn.commit()
        conn.close()

    def listar_vehiculos(self, usuario_id: Optional[int] = None) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        if usuario_id:
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM usuarios_vehiculos
                WHERE usuario_id=?
            """, (usuario_id,))
            tiene_relacion = cur.fetchone()["total"] > 0
        else:
            tiene_relacion = False

        if usuario_id and tiene_relacion:
            cur.execute("""
                SELECT v.id, v.placa, v.modelo, v.marca
                FROM vehiculos v
                JOIN usuarios_vehiculos uv ON uv.vehiculo_id = v.id
                WHERE v.activo=1 AND uv.usuario_id=?
                ORDER BY v.placa
            """, (usuario_id,))
        else:
            cur.execute("""
                SELECT id, placa, modelo, marca
                FROM vehiculos
                WHERE activo=1
                ORDER BY placa
            """)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def listar_vehiculos_disponibles(self) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT v.id, v.placa, v.modelo, v.marca
            FROM vehiculos v
            WHERE v.activo=1
              AND NOT EXISTS (
                  SELECT 1
                  FROM movimientos m
                  WHERE m.vehiculo_id = v.id
                    AND m.fecha_entrega IS NOT NULL
                    AND m.devuelto=0
                    AND NOT EXISTS (
                        SELECT 1
                        FROM movimientos_eventos me
                        WHERE me.movimiento_id = m.id
                          AND me.evento = 'RECHAZADO'
                    )
              )
            ORDER BY v.placa
        """)
        data = [dict(row) for row in cur.fetchall()]
        conn.close()
        return data

    def listar_vehiculos_prestables(self, solicitante_id: int) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT v.id, v.placa, v.modelo, v.marca,
                   u.id AS propietario_id, u.nombre AS propietario_nombre,
                   mu.fecha_entrega AS fecha_en_uso
            FROM vehiculos v
            JOIN usuarios_vehiculos uv ON uv.vehiculo_id = v.id
            JOIN usuarios u ON u.id = uv.usuario_id
            LEFT JOIN (
                SELECT vehiculo_id, MAX(fecha_entrega) AS fecha_entrega
                FROM movimientos
                WHERE fecha_entrega IS NOT NULL AND devuelto=0
                GROUP BY vehiculo_id
            ) mu ON mu.vehiculo_id = v.id
            WHERE v.activo=1
              AND u.activo=1
              AND uv.usuario_id != ?
              AND mu.fecha_entrega IS NULL
            ORDER BY u.nombre, v.placa
        """, (solicitante_id,))
        data = [dict(row) for row in cur.fetchall()]
        conn.close()
        return data

    def contar_vehiculos_disponibles(self) -> Tuple[int, int]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) AS total
            FROM vehiculos
            WHERE activo=1
        """)
        total = cur.fetchone()["total"]
        cur.execute("""
            SELECT COUNT(DISTINCT m.vehiculo_id) AS en_uso
            FROM movimientos m
            WHERE m.vehiculo_id IS NOT NULL
              AND m.fecha_entrega IS NOT NULL
              AND m.devuelto=0
              AND NOT EXISTS (
                  SELECT 1
                  FROM movimientos_eventos me
                  WHERE me.movimiento_id = m.id
                    AND me.evento = 'RECHAZADO'
              )
        """)
        en_uso = cur.fetchone()["en_uso"]
        conn.close()
        disponibles = max(total - en_uso, 0)
        return total, disponibles

    def solicitar_prestamo(
        self,
        solicitante_id: int,
        propietario_id: int,
        vehiculo_id: int,
        responsable_usuario_id: Optional[int],
        responsable_tipo: str,
        no_pasajeros: int,
        pasajeros_ids: List[int],
        fechas_solicitadas: List[str],
        ruta_destinos: List[str],
        motivo_salida: str,
        notas: Optional[str] = None,
        fecha_solicitud: Optional[str] = None,
    ) -> Tuple[bool, str]:
        if solicitante_id == propietario_id:
            return False, "No puede solicitar un prestamo de su propio vehiculo."

        requeridos = [
            ("vehiculo", vehiculo_id),
            ("responsable", responsable_usuario_id),
            ("pasajeros", no_pasajeros),
            ("fechas", fechas_solicitadas),
            ("ruta", ruta_destinos),
            ("motivo", motivo_salida),
        ]
        no_pasajeros_int = int(no_pasajeros)
        if no_pasajeros_int > 0:
            requeridos.append(("nombres_pasajeros", pasajeros_ids))
        for clave, valor in requeridos:
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                return False, f"Falta el dato de {clave}."
            if isinstance(valor, list) and not valor:
                return False, f"Falta el dato de {clave}."
        motivos_validos = {
            "Notificación de Oficio",
            "Revisión de Auditoría",
            "Entrega de Recepción",
            "Compulsas",
            "Inspección Física",
        }
        if motivo_salida not in motivos_validos:
            return False, "Motivo de salida no valido."

        fechas_limpias = []
        for fecha in fechas_solicitadas:
            fecha_txt = _parse_date(fecha)
            if not fecha_txt:
                return False, "Formato de fecha no valido para el prestamo."
            try:
                fechas_limpias.append(datetime.strptime(fecha_txt, "%Y-%m-%d").date())
            except ValueError:
                return False, "Formato de fecha no valido para el prestamo."
        if not fechas_limpias:
            return False, "Falta seleccionar al menos una fecha."
        if len(fechas_limpias) > 1:
            return False, "Solo se permite solicitar un dia por prestamo."
        hoy = date.today()
        siguientes = []
        cursor = hoy + timedelta(days=1)
        while len(siguientes) < 4:
            if cursor.weekday() < 5:
                siguientes.append(cursor)
            cursor += timedelta(days=1)
        fechas_validas = {fecha.isoformat() for fecha in siguientes}
        for fecha in fechas_limpias:
            if fecha.isoformat() not in fechas_validas:
                return False, "Las fechas deben estar dentro de los siguientes 4 dias habiles."
        fechas_txt = ",".join([fecha.isoformat() for fecha in sorted(set(fechas_limpias))])

        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT 1
            FROM usuarios_vehiculos
            WHERE usuario_id=? AND vehiculo_id=?
        """, (propietario_id, vehiculo_id))
        if not cur.fetchone():
            conn.close()
            return False, "El vehiculo no esta asignado al usuario seleccionado."

        cur.execute("""
            SELECT 1
            FROM usuarios_vehiculos
            WHERE usuario_id=? AND vehiculo_id=?
        """, (solicitante_id, vehiculo_id))
        if cur.fetchone():
            conn.close()
            return False, "El vehiculo ya esta asignado al solicitante."

        cur.execute("""
            SELECT fecha_entrega
            FROM movimientos
            WHERE vehiculo_id=?
              AND fecha_entrega IS NOT NULL
              AND devuelto=0
            ORDER BY fecha_entrega DESC
            LIMIT 1
        """, (vehiculo_id,))
        ocupacion = cur.fetchone()
        if ocupacion:
            fecha_en_uso = ocupacion["fecha_entrega"]
            dias_txt = ""
            try:
                fecha_dt = datetime.strptime(fecha_en_uso, "%Y-%m-%d").date()
                dias = max((date.today() - fecha_dt).days, 0)
                dias_txt = f" ({dias} dias)"
            except ValueError:
                dias_txt = ""
            conn.close()
            return False, f"El vehiculo esta ocupado desde {fecha_en_uso}.{dias_txt}"

        responsable_info = self._obtener_responsable(cur, responsable_tipo, responsable_usuario_id)
        if not responsable_info:
            conn.close()
            return False, "Responsable no encontrado."
        responsable_nombre, responsable_id_db = responsable_info

        pasajeros_ids = [int(pid) for pid in pasajeros_ids if pid]
        if responsable_tipo == "auditor" and responsable_usuario_id in pasajeros_ids:
            pasajeros_ids = [pid for pid in pasajeros_ids if pid != responsable_usuario_id]
            no_pasajeros_int = max(no_pasajeros_int - 1, 0)
        if no_pasajeros_int == 0:
            pasajeros_ids = []
        if pasajeros_ids:
            if len(pasajeros_ids) != len(set(pasajeros_ids)):
                conn.close()
                return False, "Pasajeros duplicados en la seleccion."
            placeholders = ",".join(["?"] * len(pasajeros_ids))
            cur.execute(f"""
                SELECT id
                FROM auditores
                WHERE id IN ({placeholders}) AND activo=1
            """, pasajeros_ids)
            pasajeros = cur.fetchall()
            if len(pasajeros) != len(set(pasajeros_ids)):
                conn.close()
                return False, "Pasajeros no encontrados."
        if no_pasajeros_int != len(pasajeros_ids):
            conn.close()
            return False, "El numero de pasajeros no coincide."

        destinos = [clave.strip().upper() for clave in ruta_destinos if clave and clave.strip()]
        if not destinos:
            conn.close()
            return False, "Falta el dato de ruta."
        placeholders_dest = ",".join(["?"] * len(destinos))
        cur.execute(f"""
            SELECT clave
            FROM entes
            WHERE clave IN ({placeholders_dest}) AND activo=1
        """, destinos)
        entes_validos = {row["clave"] for row in cur.fetchall()}
        if entes_validos != set(destinos):
            conn.close()
            return False, "Destinos no encontrados en catalogo de entes."

        cur.execute("""
            SELECT COUNT(*) AS total
            FROM prestamos_vehiculos
            WHERE solicitante_id=? AND propietario_id=? AND vehiculo_id=? AND estado='PENDIENTE'
        """, (solicitante_id, propietario_id, vehiculo_id))
        if cur.fetchone()["total"] > 0:
            conn.close()
            return False, "Ya existe una solicitud pendiente para este vehiculo."

        cur.execute("""
            INSERT INTO prestamos_vehiculos (
                solicitante_id, propietario_id, vehiculo_id, fecha_solicitud, estado, notas,
                responsable_id, responsable_nombre, no_pasajeros, pasajeros_ids, ruta_destino, motivo_salida, fechas_solicitadas
            ) VALUES (?, ?, ?, ?, 'PENDIENTE', ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            solicitante_id,
            propietario_id,
            vehiculo_id,
            _parse_date(fecha_solicitud) or _hoy_iso(),
            notas.strip() if notas else None,
            responsable_id_db,
            responsable_nombre,
            int(no_pasajeros_int),
            ",".join([str(pid) for pid in pasajeros_ids]) if pasajeros_ids else None,
            " -> ".join(destinos),
            motivo_salida.strip(),
            fechas_txt,
        ))
        conn.commit()
        conn.close()
        return True, "Solicitud de prestamo registrada."

    def listar_responsables(self) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre
            FROM responsables
            WHERE activo=1
            ORDER BY nombre
        """)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def listar_auditores(self) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre
            FROM auditores
            WHERE activo=1
            ORDER BY nombre
        """)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def crear_vehiculo(self, placa: str, marca: str, modelo: str) -> Tuple[bool, str]:
        if not placa or not marca or not modelo:
            return False, "Faltan datos requeridos."
        conn = self._connect()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO vehiculos (placa, modelo, marca, activo)
                VALUES (?, ?, ?, 1)
            """, (
                placa.strip().upper(),
                modelo.strip(),
                marca.strip().upper(),
            ))
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return False, "La placa ya existe."
        conn.close()
        return True, "Vehiculo registrado correctamente."

    # -------------------------------------------------------
    # Movimientos
    # -------------------------------------------------------
    def _obtener_responsable(self, cur, responsable_tipo: str, responsable_id: Optional[int]) -> Optional[Tuple[str, Optional[int]]]:
        if responsable_id is None:
            return None
        if responsable_tipo == "auditor":
            cur.execute("""
                SELECT id, nombre
                FROM auditores
                WHERE id=? AND activo=1
            """, (responsable_id,))
            row = cur.fetchone()
            if not row:
                return None
            return row["nombre"], None
        cur.execute("""
            SELECT id, nombre
            FROM usuarios
            WHERE id=? AND activo=1
        """, (responsable_id,))
        row = cur.fetchone()
        if not row:
            return None
        return row["nombre"], row["id"]

    def crear_movimiento(
        self,
        usuario_id: int,
        ente_clave: str,
        cantidad: int,
        receptor_nombre: str,
        firma_recepcion: str,
        observaciones: str,
        resguardante_id: Optional[int],
        vehiculo_id: int,
        responsable_usuario_id: Optional[int],
        responsable_tipo: str,
        no_pasajeros: int,
        pasajeros_ids: List[int],
        ruta_destinos: List[str],
        motivo_salida: str,
        fecha_solicitud: Optional[str] = None,
    ) -> Tuple[bool, Dict]:
        requeridos = [
            ("vehiculo", vehiculo_id),
            ("responsable", responsable_usuario_id),
            ("pasajeros", no_pasajeros),
            ("ruta", ruta_destinos),
            ("motivo", motivo_salida),
        ]
        if int(no_pasajeros) > 0:
            requeridos.append(("nombres_pasajeros", pasajeros_ids))
        for clave, valor in requeridos:
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                return False, {"mensaje": f"Falta el dato de {clave}."}
            if isinstance(valor, list) and not valor:
                return False, {"mensaje": f"Falta el dato de {clave}."}
        motivos_validos = {
            "Notificación de Oficio",
            "Revisión de Auditoría",
            "Entrega de Recepción",
            "Compulsas",
            "Inspección Física",
        }
        if motivo_salida not in motivos_validos:
            return False, {"mensaje": "Motivo de salida no valido."}

        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre
            FROM usuarios
            WHERE id=? AND activo=1
        """, (usuario_id,))
        resguardante = cur.fetchone()
        if not resguardante:
            conn.close()
            return False, {"mensaje": "Usuario no encontrado para resguardo."}

        cur.execute("""
            SELECT id, placa, modelo, marca
            FROM vehiculos
            WHERE id=? AND activo=1
        """, (vehiculo_id,))
        vehiculo = cur.fetchone()
        if not vehiculo:
            conn.close()
            return False, {"mensaje": "Vehiculo no encontrado."}

        cur.execute("""
            SELECT 1
            FROM movimientos
            WHERE vehiculo_id=?
              AND fecha_entrega IS NOT NULL
              AND devuelto=0
            LIMIT 1
        """, (vehiculo_id,))
        if cur.fetchone():
            conn.close()
            return False, {"mensaje": "La unidad ya esta en uso."}

        responsable_info = self._obtener_responsable(cur, responsable_tipo, responsable_usuario_id)
        if not responsable_info:
            conn.close()
            return False, {"mensaje": "Responsable no encontrado."}
        responsable_nombre, responsable_id_db = responsable_info

        pasajeros_ids = [int(pid) for pid in pasajeros_ids if pid]
        if responsable_tipo == "auditor" and responsable_usuario_id in pasajeros_ids:
            pasajeros_ids = [pid for pid in pasajeros_ids if pid != responsable_usuario_id]
            no_pasajeros_int = max(no_pasajeros_int - 1, 0)
        if no_pasajeros_int == 0:
            pasajeros_ids = []
        if pasajeros_ids:
            if len(pasajeros_ids) != len(set(pasajeros_ids)):
                conn.close()
                return False, {"mensaje": "Pasajeros duplicados en la seleccion."}
            placeholders = ",".join(["?"] * len(pasajeros_ids))
            cur.execute(f"""
                SELECT id, nombre
                FROM auditores
                WHERE id IN ({placeholders}) AND activo=1
                ORDER BY nombre
            """, pasajeros_ids)
            pasajeros = cur.fetchall()
            if len(pasajeros) != len(set(pasajeros_ids)):
                conn.close()
                return False, {"mensaje": "Pasajeros no encontrados."}
        if no_pasajeros_int != len(pasajeros_ids):
            conn.close()
            return False, {"mensaje": "El numero de pasajeros no coincide."}

        destinos = [clave.strip().upper() for clave in ruta_destinos if clave and clave.strip()]
        if not destinos:
            conn.close()
            return False, {"mensaje": "Falta el dato de ruta."}
        placeholders_dest = ",".join(["?"] * len(destinos))
        cur.execute(f"""
            SELECT clave
            FROM entes
            WHERE clave IN ({placeholders_dest}) AND activo=1
        """, destinos)
        entes_validos = {row["clave"] for row in cur.fetchall()}
        if entes_validos != set(destinos):
            conn.close()
            return False, {"mensaje": "Destinos no encontrados en catalogo de entes."}
        folio = self._generar_folio(cur)
        fecha_solicitud = _parse_date(fecha_solicitud) or _hoy_iso()
        receptor_nombre = receptor_nombre.strip() or resguardante["nombre"]
        ente_clave = ente_clave or destinos[0]

        cur.execute("""
            INSERT INTO movimientos (
                folio, usuario_id, ente_clave, fecha_solicitud,
                cantidad, receptor_nombre, firma_recepcion, observaciones,
                resguardante_nombre, resguardante_id, placa_unidad, marca, modelo,
                responsable_vehiculo, vehiculo_id, responsable_id,
                no_pasajeros, ruta_destino, motivo_salida
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            folio,
            usuario_id,
            ente_clave,
            fecha_solicitud,
            int(cantidad),
            receptor_nombre.strip(),
            firma_recepcion.strip() if firma_recepcion else None,
            observaciones.strip() if observaciones else None,
            resguardante["nombre"],
            resguardante["id"],
            vehiculo["placa"],
            vehiculo["marca"],
            vehiculo["modelo"],
            responsable_nombre,
            int(vehiculo_id),
            responsable_id_db,
            int(no_pasajeros_int),
            " -> ".join(destinos),
            motivo_salida.strip(),
        ))

        movimiento_id = cur.lastrowid
        if pasajeros_ids:
            cur.executemany("""
                INSERT INTO movimientos_auditores (movimiento_id, auditor_id)
                VALUES (?, ?)
            """, [(movimiento_id, pasajero_id) for pasajero_id in pasajeros_ids])
        if destinos:
            cur.executemany("""
                INSERT INTO movimientos_destinos (movimiento_id, ente_clave, orden)
                VALUES (?, ?, ?)
            """, [
                (movimiento_id, clave, orden)
                for orden, clave in enumerate(destinos, start=1)
            ])
        self._registrar_evento(cur, movimiento_id, usuario_id, "SOLICITADO", "")
        conn.commit()
        conn.close()
        return True, {"folio": folio, "movimiento_id": movimiento_id}

    def listar_movimientos(self, usuario_id: Optional[int] = None) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        q = """
            SELECT m.id, m.folio, m.ente_clave, m.fecha_solicitud,
                   m.fecha_entrega, m.fecha_devolucion, m.cantidad,
                   m.receptor_nombre, m.firma_recepcion,
                   m.devuelto, m.observaciones, m.resguardante_nombre,
                   COALESCE(m.placa_unidad, v.placa) AS placa_unidad,
                   COALESCE(m.marca, v.marca) AS marca,
                   COALESCE(m.modelo, v.modelo) AS modelo,
                   COALESCE(m.responsable_vehiculo, uresp.nombre) AS responsable_vehiculo,
                   m.vehiculo_id, m.responsable_id,
                   m.no_pasajeros,
                   COALESCE((
                       SELECT group_concat(nombre, ', ')
                       FROM (
                           SELECT a2.nombre AS nombre
                           FROM movimientos_auditores ma
                           JOIN auditores a2 ON a2.id = ma.auditor_id
                           WHERE ma.movimiento_id = m.id
                             AND a2.nombre != COALESCE(m.responsable_vehiculo, uresp.nombre)
                           ORDER BY a2.nombre
                       )
                   ), "") AS pasajeros_nombres,
                   COALESCE((
                       SELECT group_concat(destino, ' -> ')
                       FROM (
                           SELECT CASE
                               WHEN e2.clave = 'CCLET' THEN e2.clave
                               ELSE COALESCE(e2.nombre, md.ente_clave)
                           END AS destino
                           FROM movimientos_destinos md
                           LEFT JOIN entes e2 ON e2.clave = md.ente_clave
                           WHERE md.movimiento_id = m.id
                           ORDER BY md.orden
                       )
                   ), m.ruta_destino) AS ruta_destino,
                   m.motivo_salida,
                   EXISTS(
                       SELECT 1
                       FROM movimientos_eventos me
                       WHERE me.movimiento_id = m.id
                         AND me.evento = 'RECHAZADO'
                   ) AS rechazado,
                   e.nombre AS ente_nombre,
                   u.nombre AS usuario_nombre
            FROM movimientos m
            JOIN usuarios u ON u.id = m.usuario_id
            LEFT JOIN entes e ON e.clave = m.ente_clave
            LEFT JOIN vehiculos v ON v.id = m.vehiculo_id
            LEFT JOIN usuarios uresp ON uresp.id = m.responsable_id
        """
        params = ()
        if usuario_id:
            q += " WHERE m.usuario_id=?"
            params = (usuario_id,)
        q += " ORDER BY m.created_at DESC"
        cur.execute(q, params)
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        pendientes_por_vehiculo = {}
        for mov in data:
            if mov.get("vehiculo_id") and not mov.get("fecha_entrega") and not mov.get("devuelto") and not mov.get("rechazado"):
                pendientes_por_vehiculo[mov["vehiculo_id"]] = pendientes_por_vehiculo.get(mov["vehiculo_id"], 0) + 1
        for mov in data:
            vehiculo_id = mov.get("vehiculo_id")
            mov["conflicto"] = bool(vehiculo_id and pendientes_por_vehiculo.get(vehiculo_id, 0) > 1)
        return data

    def listar_agenda_semanal(self) -> Dict:
        hoy = date.today()
        lunes = hoy - timedelta(days=hoy.weekday())
        fechas = [lunes + timedelta(days=offset) for offset in range(5)]
        fechas_iso = {fecha.isoformat() for fecha in fechas}
        labels = ["Lun", "Mar", "Mie", "Jue", "Vie"]
        dias = [
            {
                "fecha": fecha.isoformat(),
                "label": f"{labels[idx]} {fecha.day:02d}/{fecha.month:02d}",
            }
            for idx, fecha in enumerate(fechas)
        ]

        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, vehiculo_id, fechas_solicitadas
            FROM prestamos_vehiculos
            WHERE estado='PENDIENTE'
        """)
        reservas_por_vehiculo = {}
        for row in cur.fetchall():
            vehiculo_id = row["vehiculo_id"]
            fechas_txt = row["fechas_solicitadas"] or ""
            fechas_reserva = {
                fecha.strip()
                for fecha in fechas_txt.split(",")
                if fecha.strip() in fechas_iso
            }
            if not fechas_reserva:
                continue
            reservas_por_vehiculo.setdefault(vehiculo_id, set()).update(fechas_reserva)

        cur.execute("""
            SELECT v.id, v.placa, v.marca, v.modelo,
                   COALESCE(GROUP_CONCAT(u.nombre, ', '), '') AS propietarios
            FROM vehiculos v
            LEFT JOIN usuarios_vehiculos uv ON uv.vehiculo_id = v.id
            LEFT JOIN usuarios u ON u.id = uv.usuario_id
            WHERE v.activo=1
            GROUP BY v.id
            ORDER BY v.placa
        """)
        vehiculos = []
        for row in cur.fetchall():
            item = dict(row)
            item["reservas"] = reservas_por_vehiculo.get(item["id"], set())
            vehiculos.append(item)
        conn.close()

        return {
            "inicio": lunes.isoformat(),
            "fin": fechas[-1].isoformat(),
            "fechas": dias,
            "vehiculos": vehiculos,
        }

    def listar_movimientos_entregados(self, usuario_id: int, fecha_iso: str) -> List[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT m.id, m.folio, m.ente_clave, m.fecha_solicitud,
                   m.fecha_entrega, m.fecha_devolucion, m.cantidad,
                   m.receptor_nombre, m.firma_recepcion,
                   m.devuelto, m.observaciones, m.resguardante_nombre,
                   COALESCE(m.placa_unidad, v.placa) AS placa_unidad,
                   COALESCE(m.marca, v.marca) AS marca,
                   COALESCE(m.modelo, v.modelo) AS modelo,
                   COALESCE(m.responsable_vehiculo, uresp.nombre) AS responsable_vehiculo,
                   m.vehiculo_id, m.responsable_id,
                   m.no_pasajeros,
                   COALESCE((
                       SELECT group_concat(nombre, ', ')
                       FROM (
                           SELECT a2.nombre AS nombre
                           FROM movimientos_auditores ma
                           JOIN auditores a2 ON a2.id = ma.auditor_id
                           WHERE ma.movimiento_id = m.id
                             AND a2.nombre != COALESCE(m.responsable_vehiculo, uresp.nombre)
                           ORDER BY a2.nombre
                       )
                   ), "") AS pasajeros_nombres,
                   COALESCE((
                       SELECT group_concat(destino, ' -> ')
                       FROM (
                           SELECT CASE
                               WHEN e2.clave = 'CCLET' THEN e2.clave
                               ELSE COALESCE(e2.nombre, md.ente_clave)
                           END AS destino
                           FROM movimientos_destinos md
                           LEFT JOIN entes e2 ON e2.clave = md.ente_clave
                           WHERE md.movimiento_id = m.id
                           ORDER BY md.orden
                       )
                   ), m.ruta_destino) AS ruta_destino,
                   m.motivo_salida,
                   e.nombre AS ente_nombre,
                   u.nombre AS usuario_nombre
            FROM movimientos m
            JOIN usuarios u ON u.id = m.usuario_id
            LEFT JOIN entes e ON e.clave = m.ente_clave
            LEFT JOIN vehiculos v ON v.id = m.vehiculo_id
            LEFT JOIN usuarios uresp ON uresp.id = m.responsable_id
            JOIN movimientos_eventos me ON me.movimiento_id = m.id
            WHERE me.evento = 'ENTREGADO'
              AND me.usuario_id = ?
              AND me.fecha = ?
            ORDER BY m.created_at DESC
        """, (usuario_id, fecha_iso))
        data = [dict(r) for r in cur.fetchall()]
        conn.close()
        return data

    def obtener_movimiento(self, movimiento_id: int) -> Optional[Dict]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT m.id, m.folio, m.ente_clave, m.fecha_solicitud,
                   m.fecha_entrega, m.fecha_devolucion, m.cantidad,
                   m.receptor_nombre, m.firma_recepcion,
                   m.devuelto, m.observaciones, m.resguardante_nombre,
                   COALESCE(m.placa_unidad, v.placa) AS placa_unidad,
                   COALESCE(m.marca, v.marca) AS marca,
                   COALESCE(m.modelo, v.modelo) AS modelo,
                   COALESCE(m.responsable_vehiculo, uresp.nombre) AS responsable_vehiculo,
                   m.vehiculo_id, m.responsable_id,
                   m.no_pasajeros,
                   COALESCE((
                       SELECT group_concat(nombre, ', ')
                       FROM (
                           SELECT a2.nombre AS nombre
                           FROM movimientos_auditores ma
                           JOIN auditores a2 ON a2.id = ma.auditor_id
                           WHERE ma.movimiento_id = m.id
                             AND a2.nombre != COALESCE(m.responsable_vehiculo, uresp.nombre)
                           ORDER BY a2.nombre
                       )
                   ), "") AS pasajeros_nombres,
                   COALESCE((
                       SELECT group_concat(destino, ' -> ')
                       FROM (
                           SELECT CASE
                               WHEN e2.clave = 'CCLET' THEN e2.clave
                               ELSE COALESCE(e2.nombre, md.ente_clave)
                           END AS destino
                           FROM movimientos_destinos md
                           LEFT JOIN entes e2 ON e2.clave = md.ente_clave
                           WHERE md.movimiento_id = m.id
                           ORDER BY md.orden
                       )
                   ), m.ruta_destino) AS ruta_destino,
                   m.motivo_salida,
                   u.nombre AS usuario_nombre,
                   e.nombre AS ente_nombre
            FROM movimientos m
            JOIN usuarios u ON u.id = m.usuario_id
            LEFT JOIN entes e ON e.clave = m.ente_clave
            LEFT JOIN vehiculos v ON v.id = m.vehiculo_id
            LEFT JOIN usuarios uresp ON uresp.id = m.responsable_id
            WHERE m.id=?
            LIMIT 1
        """, (movimiento_id,))
        row = cur.fetchone()
        conn.close()
        return dict(row) if row else None

    def marcar_entregado(self, movimiento_id: int, usuario_id: int) -> Tuple[bool, str]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, fecha_entrega, devuelto
            FROM movimientos
            WHERE id=?
        """, (movimiento_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return False, "Movimiento no encontrado."
        cur.execute("""
            SELECT 1
            FROM movimientos_eventos
            WHERE movimiento_id=?
              AND evento='RECHAZADO'
            LIMIT 1
        """, (movimiento_id,))
        if cur.fetchone():
            conn.close()
            return False, "El movimiento fue rechazado."
        if row["fecha_entrega"]:
            conn.close()
            return False, "El movimiento ya fue entregado."
        cur.execute("""
            UPDATE movimientos
            SET fecha_entrega=?
            WHERE id=?
        """, (_hoy_iso(), movimiento_id))
        self._registrar_evento(cur, movimiento_id, usuario_id, "ENTREGADO", "")
        conn.commit()
        conn.close()
        return True, "Entrega registrada."

    def marcar_rechazado(self, movimiento_id: int, usuario_id: int) -> Tuple[bool, str]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, fecha_entrega, devuelto
            FROM movimientos
            WHERE id=?
        """, (movimiento_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return False, "Movimiento no encontrado."
        if row["fecha_entrega"] or row["devuelto"]:
            conn.close()
            return False, "No se puede rechazar un movimiento entregado."
        cur.execute("""
            SELECT 1
            FROM movimientos_eventos
            WHERE movimiento_id=?
              AND evento='RECHAZADO'
            LIMIT 1
        """, (movimiento_id,))
        if cur.fetchone():
            conn.close()
            return False, "El movimiento ya fue rechazado."
        self._registrar_evento(cur, movimiento_id, usuario_id, "RECHAZADO", "")
        conn.commit()
        conn.close()
        return True, "Movimiento rechazado."

    def marcar_devuelto(self, movimiento_id: int, usuario_id: int) -> Tuple[bool, str]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, fecha_entrega, devuelto
            FROM movimientos
            WHERE id=?
        """, (movimiento_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return False, "Movimiento no encontrado."
        if row["devuelto"]:
            conn.close()
            return False, "El movimiento ya está devuelto."
        if not row["fecha_entrega"]:
            conn.close()
            return False, "No se puede devolver sin entrega registrada."
        cur.execute("""
            UPDATE movimientos
            SET fecha_devolucion=?, devuelto=1
            WHERE id=?
        """, (_hoy_iso(), movimiento_id))
        self._registrar_evento(cur, movimiento_id, usuario_id, "DEVUELTO", "")
        conn.commit()
        conn.close()
        return True, "Devolución registrada."

    def movimientos_con_alerta(self, movimientos: List[Dict], dias_alerta: int) -> List[MovimientoRow]:
        rows = []
        for mov in movimientos:
            alerta = False
            if mov.get("fecha_entrega") and not mov.get("devuelto"):
                try:
                    fecha_entrega = datetime.strptime(mov["fecha_entrega"], "%Y-%m-%d").date()
                    alerta = date.today() - fecha_entrega >= timedelta(days=dias_alerta)
                except ValueError:
                    alerta = False
            rows.append(MovimientoRow(data=mov, alerta=alerta))
        return rows

    def _registrar_evento(self, cur, movimiento_id: int, usuario_id: int, evento: str, notas: str):
        cur.execute("""
            INSERT INTO movimientos_eventos (movimiento_id, usuario_id, evento, fecha, notas)
            VALUES (?, ?, ?, ?, ?)
        """, (movimiento_id, usuario_id, evento, _hoy_iso(), notas))

    def _generar_folio(self, cur) -> str:
        hoy = datetime.now().strftime("%Y%m%d")
        cur.execute("""
            SELECT COUNT(*) FROM movimientos
            WHERE fecha_solicitud=?
        """, (_hoy_iso(),))
        contador = cur.fetchone()[0] + 1
        return f"ACTA-{hoy}-{contador:04d}"
